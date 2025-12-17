#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 엑셀 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "API 명세"

# 스타일 정의
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 컬럼 너비 설정
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 30

# 제목 행
ws['A1'] = "EatWise API 명세 (2025-11-25)"
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws.merge_cells('A1:F1')
ws['A1'].alignment = center_alignment

# 헤더 행
headers = ["도메인", "HTTP 메서드", "엔드포인트", "설명", "요청(Request)", "응답(Response)"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# API 데이터
api_data = [
    # User API
    ("User", "POST", "/api/users/signup", "새 사용자 생성(회원가입)",
     "username, password, email, age, gender, height, weight",
     "UserResponse: userId, username, email, age, gender, height, weight"),

    ("User", "POST", "/api/users/login", "로그인 (이메일/비밀번호)",
     "email, password",
     "UserResponse: 사용자 정보"),

    ("User", "PUT", "/api/users/{userId}", "회원 정보 수정",
     "password, age, height, weight",
     "UserResponse: 수정된 사용자"),

    ("User", "GET", "/api/users/{userId}", "특정 사용자 조회",
     "경로변수: userId",
     "UserResponse"),

    # Auth API
    ("Auth", "POST", "/api/auth", "인증 레코드 생성 (이메일 인증코드 저장)",
     "userId, email, verificationCode, expiresAt(LocalDateTime), verified",
     "AuthResponse"),

    ("Auth", "POST", "/api/auth/send-code", "이메일로 인증코드 전송",
     "?email={email}",
     "AuthResponse"),

    ("Auth", "POST", "/api/auth/verify", "인증 코드 검증",
     "code, email 등",
     "AuthResponse"),

    ("Auth", "GET", "/api/auth/{authId}", "인증 레코드 조회",
     "경로변수: authId",
     "AuthResponse"),

    ("Auth", "GET", "/api/auth/user/{userId}", "사용자 인증 레코드 조회",
     "경로변수: userId",
     "List<AuthResponse>"),

    ("Auth", "GET", "/api/auth/email/{email}/latest", "최신 인증 레코드 조회",
     "경로변수: email",
     "AuthResponse"),

    # Food API
    ("Food", "POST", "/api/foods", "음식(식품) 등록",
     "foodName, category, calories, carbs, protein, fat, imageUrl",
     "FoodResponse: foodId, foodName, category, 영양정보"),

    ("Food", "GET", "/api/foods/{foodId}", "특정 음식 조회",
     "경로변수: foodId",
     "FoodResponse"),

    ("Food", "GET", "/api/foods", "전체 음식 목록 조회",
     "쿼리: 없음",
     "List<FoodResponse>"),

    ("Food", "GET", "/api/foods/category/{category}", "카테고리별 음식 조회",
     "경로변수: category",
     "List<FoodResponse>"),

    ("Food", "GET", "/api/foods/search", "음식명 검색",
     "?keyword={keyword}",
     "List<FoodResponse>"),

    ("Food", "POST", "/api/foods/extract-from-ocr", "OCR 텍스트에서 음식명 추출",
     "ocrText (String)",
     "List<FoodResponse> (추출 결과)"),

    # MealRecord API
    ("MealRecord", "POST", "/api/meal-records", "식사 기록 생성",
     "userId, foodId, mealType, intakeTime(HH:mm:ss), intakeDate(yyyy-MM-dd), quantity",
     "MealRecordResponse: recordId, userId, foodId, 식사정보, totalCalories"),

    ("MealRecord", "GET", "/api/meal-records/{recordId}", "단일 식사 기록 조회",
     "경로변수: recordId",
     "MealRecordResponse"),

    ("MealRecord", "GET", "/api/meal-records/user/{userId}", "사용자의 모든 식사 기록 조회",
     "경로변수: userId",
     "List<MealRecordResponse>"),

    ("MealRecord", "GET", "/api/meal-records/user/{userId}/date/{date}", "특정 날짜 식사 기록 조회",
     "경로변수: userId, date(yyyy-MM-dd)",
     "List<MealRecordResponse>"),

    ("MealRecord", "GET", "/api/meal-records/user/{userId}/date-range", "기간별 식사 기록 조회(달력용)",
     "경로: userId, ?startDate=yyyy-MM-dd&endDate=yyyy-MM-dd",
     "List<MealRecordResponse>"),

    ("MealRecord", "PUT", "/api/meal-records/{recordId}", "식사 기록 수정",
     "mealType, intakeTime, intakeDate, quantity",
     "MealRecordResponse"),

    ("MealRecord", "DELETE", "/api/meal-records/{recordId}", "식사 기록 삭제",
     "경로변수: recordId",
     "없음 (204 No Content)"),

    ("MealRecord", "POST", "/api/meal-records/ocr", "OCR 영수증 처리",
     "userId, imageBase64, intakeDate, mealType, recognizedText, foods[]",
     "OcrResponse: message, foods[]"),

    # Goal API
    ("Goal", "POST", "/api/goals", "사용자 목표 생성(체중 관리)",
     "userId, goalType(GAIN/MAINTAIN/LOSE), dailyCalorieTarget, carbRatio, proteinRatio, fatRatio, startDate, endDate",
     "GoalResponse: goalId, 목표정보"),

    ("Goal", "GET", "/api/goals/{goalId}", "목표 조회",
     "경로변수: goalId",
     "GoalResponse"),

    ("Goal", "GET", "/api/goals/user/{userId}", "사용자의 모든 목표 조회",
     "경로변수: userId",
     "List<GoalResponse>"),

    ("Goal", "GET", "/api/goals/user/{userId}/current", "현재 활성 목표 조회",
     "경로변수: userId",
     "GoalResponse"),

    ("Goal", "GET", "/api/goals/user/{userId}/active", "활성 목표들 조회",
     "경로변수: userId",
     "List<GoalResponse>"),

    ("Goal", "PUT", "/api/goals/{goalId}", "목표 수정",
     "goalType, dailyCalorieTarget, 영양소 비율 등",
     "GoalResponse"),

    # Notification API
    ("Notification", "POST", "/api/notifications", "알림 생성",
     "userId, type, message, createdAt(LocalDateTime), readStatus",
     "NotificationResponse: notificationId, 알림정보"),

    ("Notification", "GET", "/api/notifications/{notificationId}", "알림 조회",
     "경로변수: notificationId",
     "NotificationResponse"),

    ("Notification", "GET", "/api/notifications/user/{userId}", "사용자 알림 조회",
     "경로변수: userId",
     "List<NotificationResponse>"),

    ("Notification", "GET", "/api/notifications/user/{userId}/unread", "읽지 않은 알림 조회",
     "경로변수: userId",
     "List<NotificationResponse>"),

    ("Notification", "GET", "/api/notifications/user/{userId}/unread/count", "읽지 않은 알림 개수",
     "경로변수: userId",
     "Long (개수)"),

    ("Notification", "PATCH", "/api/notifications/{notificationId}/read", "알림 읽음 표시",
     "경로변수: notificationId",
     "NotificationResponse"),

    ("Notification", "PATCH", "/api/notifications/user/{userId}/read-all", "모든 알림 읽음 처리",
     "경로변수: userId",
     "없음 (200 OK)"),

    # RecommendedFood API
    ("RecommendedFood", "POST", "/api/recommended-foods", "추천 음식 생성",
     "userId, foodId, reason, createdAt(LocalDateTime)",
     "RecommendedFoodResponse: recId, 추천정보"),

    ("RecommendedFood", "GET", "/api/recommended-foods/{recId}", "추천 음식 조회",
     "경로변수: recId",
     "RecommendedFoodResponse"),

    ("RecommendedFood", "GET", "/api/recommended-foods/user/{userId}", "사용자 추천 음식 조회",
     "경로변수: userId",
     "List<RecommendedFoodResponse>"),

    ("RecommendedFood", "GET", "/api/recommended-foods/user/{userId}/recent", "최근 추천 음식 조회",
     "경로: userId, ?days=7(기본값)",
     "List<RecommendedFoodResponse>"),

    ("RecommendedFood", "PUT", "/api/recommended-foods/{recId}", "추천 음식 수정",
     "reason, createdAt 등",
     "RecommendedFoodResponse"),

    # Restaurant API
    ("Restaurant", "POST", "/api/restaurants", "음식점 등록",
     "name, address, latitude, longitude 등",
     "RestaurantResponse: restaurantId, 음식점정보"),

    ("Restaurant", "GET", "/api/restaurants/{restaurantId}", "음식점 조회",
     "경로변수: restaurantId",
     "RestaurantResponse"),

    ("Restaurant", "GET", "/api/restaurants", "모든 음식점 조회",
     "쿼리: 없음",
     "List<RestaurantResponse>"),

    ("Restaurant", "GET", "/api/restaurants/search/name", "음식점명 검색",
     "?keyword={keyword}",
     "List<RestaurantResponse>"),

    ("Restaurant", "GET", "/api/restaurants/search/address", "주소로 검색",
     "?address={address}",
     "List<RestaurantResponse>"),

    ("Restaurant", "GET", "/api/restaurants/nearby", "근처 음식점 조회",
     "?latitude=&longitude=&radius=5.0(km)",
     "List<RestaurantResponse>"),

    ("Restaurant", "PUT", "/api/restaurants/{restaurantId}", "음식점 정보 수정",
     "name, address, 좌표 등",
     "RestaurantResponse"),

    # Report API
    ("Report", "POST", "/api/reports", "보고서(리포트) 생성",
     "userId, reportDate 등",
     "ReportResponse: reportId, 보고서정보"),

    ("Report", "GET", "/api/reports/{reportId}", "보고서 조회",
     "경로변수: reportId",
     "ReportResponse"),

    ("Report", "GET", "/api/reports/user/{userId}", "사용자 보고서 조회",
     "경로변수: userId",
     "List<ReportResponse>"),

    ("Report", "GET", "/api/reports/user/{userId}/date/{reportDate}", "특정 날짜 보고서 조회",
     "경로: userId, reportDate(yyyy-MM-dd)",
     "ReportResponse"),

    ("Report", "GET", "/api/reports/user/{userId}/date-range", "기간별 보고서 조회",
     "경로: userId, ?startDate=yyyy-MM-dd&endDate=yyyy-MM-dd",
     "List<ReportResponse>"),

    ("Report", "PUT", "/api/reports/{reportId}", "보고서 수정",
     "reportDate 등",
     "ReportResponse"),
]

# 데이터 입력
row = 4
for domain, method, endpoint, description, request, response in api_data:
    ws.cell(row=row, column=1).value = domain
    ws.cell(row=row, column=2).value = method
    ws.cell(row=row, column=3).value = endpoint
    ws.cell(row=row, column=4).value = description
    ws.cell(row=row, column=5).value = request
    ws.cell(row=row, column=6).value = response

    # 스타일 적용
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = left_alignment

        # HTTP 메서드별 색상
        if col == 2:
            if method == "POST":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif method == "GET":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif method == "PUT":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif method == "PATCH":
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif method == "DELETE":
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            cell.font = Font(bold=True)

    row += 1

# 행 높이 자동 조정
ws.row_dimensions[3].height = 25
for r in range(4, row):
    ws.row_dimensions[r].height = None  # 자동 높이

# 시트 2: 주의사항
ws2 = wb.create_sheet("주의사항 & 설정")

# 시트2 헤더
ws2['A1'] = "주의사항 & 설정 정보"
ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws2.merge_cells('A1:B1')
ws2['A1'].alignment = center_alignment

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 80

# 주의사항 데이터
notes = [
    ("날짜 포맷 (LocalDate)", "yyyy-MM-dd (예: 2025-11-21)"),
    ("시간 포맷 (LocalDateTime)", "yyyy-MM-dd'T'HH:mm:ss 또는 변환된 문자열"),
    ("식사시간 포맷 (intakeTime)", "HH:mm:ss (예: 14:30:00)"),
    ("식사날짜 포맷 (intakeDate)", "yyyy-MM-dd (예: 2025-11-21)"),
    ("Jackson 라이브러리", "jackson-datatype-jsr310 모듈이 필요함 (LocalDate/LocalDateTime 직렬화용)"),
    ("기본 오류 메시지", "Type definition error: [simple type, class java.time.LocalDate] → Jackson 모듈 추가 필요"),
    ("Bean 충돌 에러", "objectMapper 중복 정의 시 발생 → WebConfig/JacksonConfig 중 하나 제거"),
    ("Ambiguous Mapping", "중복된 @RequestMapping 경로가 있으면 발생 → 중복 파일/클래스 제거"),
    ("OCR/LLM 호출", "외부 API(OpenAI, HuggingFace) 사용 시 401/404/410 에러 가능 → API 키 또는 엔드포인트 확인"),
    ("식사기록 조회(달력용)", "GET /api/meal-records/user/{userId}/date-range 사용"),
    ("HTTP 상태코드", "201: 생성, 200: 성공, 204: 삭제성공, 4xx: 클라이언트오류, 5xx: 서버오류"),
    ("Null DateTime 처리", "LocalDateTime은 가능하면 프론트엔드에서 문자열로 변환 후 전송"),
]

row = 3
for topic, description in notes:
    ws2.cell(row=row, column=1).value = topic
    ws2.cell(row=row, column=2).value = description
    ws2.cell(row=row, column=1).font = Font(bold=True, size=10)
    ws2.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws2.cell(row=row, column=2).alignment = left_alignment
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=2).border = thin_border
    ws2.row_dimensions[row].height = 25
    row += 1

# 파일 저장
output_path = "EatWise_API_Specification.xlsx"
wb.save(output_path)
print(f"✅ 엑셀 파일이 생성되었습니다: {output_path}")
print(f"📊 총 {len(api_data)}개의 API 엔드포인트가 포함되었습니다.")

